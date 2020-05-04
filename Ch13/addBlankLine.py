# AutomateTheboringStuff Ch13: Blank Row Inserter

# Create a program blankRowInserter.py that takes two integers and a filename
# string as command line arguments. Let’s call the first integer N and the
# second integer M. Starting at row N, the program should insert M blank rows
# into the spreadsheet.

import openpyxl
import sys

# get input from command line
if len(sys.argv) != 4:
    print("ERR1: Invalid arguments: blankRowInserter <N> <M> <filename>")
    sys.exit()
try:
    N = int(sys.argv[1])
    M = int(sys.argv[2])
    filename = sys.argv[3]
except ValueError:
    print("ERR2: Invalid arguments: blankRowInserter <N> <M> <filename>")
    sys.exit()

# prepare spreadsheet
try:
    sourceWb = openpyxl.load_workbook(filename)
except FileNotFoundError:
    print(f"ERR3: Cannot open {filename}")
    sys.exit()
destWb = openpyxl.Workbook()
destSheet = destWb.active
sourceSheet = sourceWb.active

# copy until N row
for rowIndex in range(1, N):
    for colIndex in range(1, sourceSheet.max_column + 1):
        destSheet.cell(column=colIndex, row=rowIndex).value = \
        sourceSheet.cell(column=colIndex, row=rowIndex).value

# copy from N+M row on
for rowIndex in range(N,sourceSheet.max_row + 1):
    for colIndex in range(1, sourceSheet.max_column + 1):
        destSheet.cell(column=colIndex, row=rowIndex + M).value = \
        sourceSheet.cell(column=colIndex, row=rowIndex).value
# save
destWb.save("new_" + filename)
