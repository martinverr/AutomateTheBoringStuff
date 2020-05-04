
# AutomateTheboringStuff Ch13: Spreadsheet Cell Inverter

# Write a program to invert the row and column of the cells in the spreadsheet.
# For example, the value at row 5, column 3 will be at row 3, column 5 (and
# vice versa). This should be done for all cells in the spreadsheet.

import openpyxl


def cellInverter(source, dest=None):
    # return -1 if error opening source

    # prepare spreadsheet
    try:
        sourceWb = openpyxl.load_workbook(source)
    except FileNotFoundError:
        print(f"ERR: Cannot open {source}")
        return -1
    destWb = openpyxl.Workbook()
    destSheet = destWb.active
    sourceSheet = sourceWb.active

    # copy until N row
    for rowIndex in range(1, sourceSheet.max_row + 1):
        for colIndex in range(1, sourceSheet.max_column + 1):
            destSheet.cell(column=rowIndex, row=colIndex).value = \
                sourceSheet.cell(column=colIndex, row=rowIndex).value

    # save
    if dest is None:
        dest = source
    destWb.save(dest)

# TEST___________________

print("testing with a inexsistent file")
cellInverter("notfound.xlsx")
print("testing with test.xlsx")
cellInverter("test.xlsx","testInverted.xlsx")
print("test inverting two times a file")
cellInverter("test.xlsx")
cellInverter("test.xlsx")
