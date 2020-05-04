# AutomateTheBoringStuff Ch13: Spreadsheet to Text Files

# Write a program that performs the tasks of the previous program in reverse
# order: the program should open a spreadsheet and write the cells of column A
# into one text file, the cells of column B into another text file, and so on.

import openpyxl


def xlsx2txt(spreadsheet):
    # check if exsist
    try:
        wb = openpyxl.load_workbook(spreadsheet)
    except FileNotFoundError:
        print("The spreadsheet doesn't exsist")
        return

    sheet = wb.active

    i = 1  # coulumn index = txt file number
    while i < sheet.max_column + 1:

        # get the text from the rows
        text = ""
        for rowIndex in range(1, sheet.max_row + 1):
            if sheet.cell(row=rowIndex, column=i).value is not None:
                text = text + sheet.cell(row=rowIndex, column=i).value

        # copy text to file
        f = open(f"{spreadsheet[:-4]}{i}.txt", "w")
        f.write(text)
        f.close()

        i += 1

# MAIN__________________________


xlsx2txt("output.xlsx")
