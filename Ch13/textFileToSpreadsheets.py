import openpyxl


def txt2xlsx(txtList, outputName="output.xlsx"):
    # checks
    if txtList is None or txtList == []:
        print("The list of txt files is empty")
    for file in txtList:
        if not file.endswith(".txt"):
            print("The parameter list contains non-txt files")

    wb = openpyxl.Workbook()
    sheet = wb.active

    i = 0  # coulumn-1 index and txtList index
    while i < len(txtList):
        # read txt file lines
        f = open(txtList[i], "r")
        lines = f.readlines()
        f.close()

        colLenght = 0
        for rowIndex in range(len(lines)):
            # Copy lines in each row, while updating column width
            sheet.cell(row=rowIndex+1, column=i+1).value = lines[rowIndex]
            if len(lines[rowIndex]) > colLenght:
                colLenght = len(lines[rowIndex])
        # set column width
        sheet.column_dimensions[openpyxl.utils.get_column_letter(i+1)].width = colLenght

        i += 1

    wb.save(outputName)

# MAIN__________________________


files = ["test1.txt", "test2.txt"]
txt2xlsx(files)
