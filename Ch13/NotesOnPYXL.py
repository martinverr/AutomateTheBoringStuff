import openpyxl

wb = openpyxl.load_workbook('example.xlsx')
wb.sheetnames  # The workbook's sheets' names.

sheet = wb['Sheet1']  # Get a sheet from the workbook.
sheet.title  # Get the sheet's title as a string.
anotherSheet = wb.active  # Get the active sheet.


sheet['A1'].value  # Get the value from the cell.
sheet.max_row  # Get the highest row number.
sheet.max_column  # Get the highest column number.
sheet.cell(row=1, column=2)
for i in range(1, 8, 2):  # Go through every other row:
    print(i, sheet.cell(row=i, column=2).value)

cell = sheet['A1']  # Get a cell from the sheet.
print('Row %s, Column %s is %s' % (cell.row, cell.column, cell.value))
print('Cell %s is %s' % (cell.coordinate, cell.value))


openpyxl.utils.get_column_letter(1)
openpyxl.utils.column_index_from_string('A')

list(sheet['A1':'C3'])  # Get all cells from A1 to C3.
for rowOfCellObjects in sheet['A1':'C3']:
    for cellObj in rowOfCellObjects:
        print(cellObj.coordinate, cellObj.value)
    print('--- END OF ROW ---')
list(sheet.columns)[1]  # Get second column's cells.

""" SUMMARY
Import the openpyxl module.
Call the openpyxl.load_workbook() function.
Get a Workbook object.
Use the active or sheetnames attributes.
Get a Worksheet object.
Use indexing or the cell() sheet method with row and column keyword arguments.
Get a Cell object.
Read the Cell object’s value attribute.
"""
