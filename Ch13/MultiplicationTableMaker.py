# Multiplication Table Maker

# Create a program multiplicationTable.py that takes a number N from the
# command line and creates an N×N multiplication table in an Excel spreadsheet.

import sys
import openpyxl

# Get N from sys.argv
if len(sys.argv) != 2:  # program name + one argument
    print("ERR: Invalid number of arguments input")
    sys.exit()
try:
    N = int(sys.argv[1])
    if N < 0:
        print("ERR: Invalid argument type(not positive")
        sys.exit()
except ValueError:
    print("ERR: Invalid argument type(not an integer)")
    sys.exit()

# Prepare spreadsheet
wb = openpyxl.Workbook()
sheet = wb.active
sheet.name = f"{N} * {N} multiplicationTable"
boldFont = openpyxl.styles.Font(b=True)

# Set headers
for colIndex in range(1, N+1):
    sheet[openpyxl.utils.get_column_letter(colIndex+1)+'1'].value = colIndex
    sheet[openpyxl.utils.get_column_letter(colIndex+1)+'1'].font = boldFont
for rowIndex in range(1, N+1):
    sheet.cell(row=rowIndex+1, column=1).value = rowIndex
    sheet.cell(row=rowIndex+1, column=1).font = boldFont


# Set the values of multiplicationTable
for rows in range(1, N+1):
    for cols in range(1, N+1):
        sheet.cell(row=rows+1, column=cols+1).value = rows * cols
wb.save(f"multiplicationTable{N}.xlsx")
