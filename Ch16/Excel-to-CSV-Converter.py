import os
import openpyxl, csv

for excelFile in os.listdir('.'):
    # Skip non-xlsx files, load the workbook object.
    if not excelFile.endswith('.xlsx'):
        continue
    wb = openpyxl.load_workbook(excelFile)

    for sheetName in wb.get_sheet_names():
        sheet = wb.get_sheet_by_name(sheetName)

        # Create the CSV filename from the Excel filename and sheet title.
        csvFile = open(filename[:-5]+'_'+sheetName, "w", newline='')

        # Create the csv.writer object for this CSV file.
        csvObj = csv.writer(csvFile)

        # Loop through every row in the sheet.
        for rowNum in range(1, sheet.max_row + 1):
            rowData = []    # append each cell to this list
            # Loop through each cell in the row.
            for colNum in range(1, sheet.max_column + 1):
                # Append each cell's data to rowData.
                rowData.append(sheet.cell(row=rowNum, col=colNum).value)

            # Write the rowData list to the CSV file.
            csv.writerow(rowData)

        csvFile.close()
